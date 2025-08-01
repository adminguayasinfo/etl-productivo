"""
Extractor para datos de fertilizantes desde archivo Excel.
"""
import openpyxl
from datetime import datetime, date
from typing import Iterator, List, Dict, Any, Optional
from loguru import logger


class FertilizantesExcelExtractor:
    """Extrae datos de fertilizantes desde archivo Excel."""
    
    def __init__(self, excel_path: str):
        """
        Inicializa el extractor.
        
        Args:
            excel_path: Ruta al archivo Excel
        """
        self.excel_path = excel_path
        self.sheet_name = 'FERTILIZANTES'
        
        # Mapeo de headers del Excel a campos del modelo
        self.field_mapping = {
            'FECHA DE ENTREGA': 'fecha_entrega',
            'ASOCIACIONES': 'asociaciones',
            'APELLIDOS Y NOMBRES': 'nombres_apellidos',
            'CEDULA': 'cedula',
            'TELEFONO': 'telefono',
            'GENERO': 'genero',
            'EDAD': 'edad',
            'CANTON': 'canton',
            'PARROQUIA': 'parroquia',
            'RECINTO, COMUNA O SECTOR': 'recinto',
            'X': 'coord_x',
            'Y': 'coord_y',
            'HECTAREAS': 'hectareas',
            'FERTILIZANTE NITROGENADO': 'fertilizante_nitrogenado',
            '(N-P-K) + ELEMENTOS MENORES': 'npk_elementos_menores',
            'ORGANICO FOLIAR': 'organico_foliar',
            'CULTIVO': 'cultivo',
            'Precio_Kit': 'precio_kit',
            'LUGAR DE ENTREGA': 'lugar_entrega',
            'OBSERVACION': 'observacion',
            'AÑO': 'anio'
        }
    
    def safe_int(self, value) -> Optional[int]:
        """Convierte valor a entero de forma segura."""
        if value is None:
            return None
        
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
        
        try:
            return int(float(str_val))
        except (ValueError, TypeError):
            return None
    
    def safe_float(self, value) -> Optional[float]:
        """Convierte valor a float de forma segura."""
        if value is None:
            return None
        
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
        
        try:
            return float(str_val)
        except (ValueError, TypeError):
            return None
    
    def safe_date(self, value) -> Optional[date]:
        """Convierte valor a fecha de forma segura."""
        if value is None:
            return None
        
        # Si ya es una fecha/datetime
        if isinstance(value, (date, datetime)):
            if isinstance(value, datetime):
                return value.date()
            return value
        
        # Si es string, intentar parsear
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
        
        try:
            # Intentar diferentes formatos
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(str_val, fmt).date()
                except ValueError:
                    continue
            return None
        except Exception:
            return None
    
    def safe_string(self, value) -> Optional[str]:
        """Convierte valor a string de forma segura."""
        if value is None:
            return None
        
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
        
        return str_val
    
    def extract_batches(self, batch_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
        """
        Extrae datos en lotes desde Excel.
        
        Args:
            batch_size: Tamaño de lote
            
        Yields:
            Lista de diccionarios con los datos de cada lote
        """
        logger.info(f"Extrayendo datos en lotes de {batch_size} registros")
        
        # Cargar workbook
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[self.sheet_name]
        
        # Obtener headers
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                break
        
        logger.info(f"Headers encontrados: {len(headers)} columnas")
        
        # Procesar datos en lotes
        batch = []
        total_rows = 0
        
        for row_num in range(2, sheet.max_row + 1):
            # Leer fila
            row_data = {}
            has_data = False
            
            for col_num, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                
                # Mapear header a campo del modelo
                if header in self.field_mapping:
                    field_name = self.field_mapping[header]
                    
                    # Aplicar conversión según el tipo de campo
                    if field_name == 'fecha_entrega':
                        row_data[field_name] = self.safe_date(cell_value)
                    elif field_name in ['edad', 'fertilizante_nitrogenado', 'npk_elementos_menores', 'organico_foliar', 'anio']:
                        row_data[field_name] = self.safe_int(cell_value)
                    elif field_name in ['hectareas', 'precio_kit']:
                        row_data[field_name] = self.safe_float(cell_value)
                    else:
                        row_data[field_name] = self.safe_string(cell_value)
            
            # Solo agregar si la fila tiene datos
            if has_data:
                batch.append(row_data)
                total_rows += 1
            
            # Enviar lote cuando esté lleno
            if len(batch) >= batch_size:
                logger.info(f"Enviando lote con {len(batch)} registros (total procesadas: {total_rows} filas)")
                yield batch
                batch = []
        
        # Enviar último lote si tiene datos
        if batch:
            logger.info(f"Enviando último lote con {len(batch)} registros (total procesadas: {total_rows} filas)")
            yield batch
        
        workbook.close()
        logger.info(f"Extracción completada: {total_rows} registros totales")
    
    def get_total_rows(self) -> int:
        """Obtiene el total de filas con datos en el Excel."""
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[self.sheet_name]
        
        # Contar filas con datos (excluyendo header)
        count = 0
        for row_num in range(2, sheet.max_row + 1):
            # Verificar si la fila tiene al menos un dato
            has_data = False
            for col_num in range(1, min(22, sheet.max_column + 1)):  # Máximo 21 columnas esperadas
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            
            if has_data:
                count += 1
        
        workbook.close()
        return count