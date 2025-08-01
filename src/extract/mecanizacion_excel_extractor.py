"""
Extractor para datos de mecanización desde archivo Excel.
"""
import openpyxl
from datetime import datetime, date
from typing import Iterator, List, Dict, Any, Optional
from loguru import logger


class MecanizacionExcelExtractor:
    """Extrae datos de mecanización desde archivo Excel."""
    
    def __init__(self, excel_path: str):
        """
        Inicializa el extractor.
        
        Args:
            excel_path: Ruta al archivo Excel
        """
        self.excel_path = excel_path
        self.sheet_name = 'MECANIZACIÓN'
        
        # Mapeo de headers del Excel a campos del modelo
        self.field_mapping = {
            'APELLIDOS Y NOMBRES': 'nombres_apellidos',
            'CÉDULA DE IDENTIDAD': 'cedula',
            'NÚMERO DE TELÉFONO': 'telefono',
            'Género': 'genero',
            'EDAD': 'edad',
            'CANTON': 'canton',
            'AGRUPACIÓN': 'agrupacion',
            'RECINTO, COMUNA O SECTOR': 'recinto',
            'X': 'coord_x',
            'Y': 'coord_y',
            'HECTÁREAS BENEFICIADAS': 'hectareas_beneficiadas',
            'CULTIVO': 'cultivo',
            'ESTADO': 'estado',
            'COMENTARIO': 'comentario',
            'CU-HA': 'cu_ha',
            'INVERSION': 'inversion',
            'AÑO': 'anio'
        }
    
    def safe_int(self, value) -> Optional[int]:
        """Convierte valor a entero de forma segura."""
        if value is None:
            return None
        
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
        
        try:
            return int(float(str_val))
        except (ValueError, TypeError):
            return None
    
    def safe_float(self, value) -> Optional[float]:
        """Convierte valor a float de forma segura."""
        if value is None:
            return None
        
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
            
        # Si es una fórmula de Excel (empieza con =), retornar None
        if str_val.startswith('='):
            return None
        
        try:
            return float(str_val)
        except (ValueError, TypeError):
            return None
    
    def safe_string(self, value) -> Optional[str]:
        """Convierte valor a string de forma segura."""
        if value is None:
            return None
        
        str_val = str(value).strip()
        if str_val in ['', ' ', 'nan', 'NaN', 'None']:
            return None
            
        # Si es una fórmula de Excel (empieza con =), retornar None
        if str_val.startswith('='):
            return None
        
        return str_val
    
    def extract_batches(self, batch_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
        """
        Extrae datos en lotes desde Excel.
        
        Args:
            batch_size: Tamaño de lote
            
        Yields:
            Lista de diccionarios con los datos de cada lote
        """
        logger.info(f"Extrayendo datos en lotes de {batch_size} registros")
        
        # Cargar workbook
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[self.sheet_name]
        
        # Obtener headers
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                break
        
        logger.info(f"Headers encontrados: {len(headers)} columnas")
        
        # Procesar datos en lotes
        batch = []
        total_rows = 0
        
        for row_num in range(2, sheet.max_row + 1):
            # Leer fila
            row_data = {}
            has_data = False
            
            for col_num, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                
                # Mapear header a campo del modelo
                if header in self.field_mapping:
                    field_name = self.field_mapping[header]
                    
                    # Aplicar conversión según el tipo de campo
                    if field_name in ['edad', 'anio']:
                        row_data[field_name] = self.safe_int(cell_value)
                    elif field_name in ['hectareas_beneficiadas', 'cu_ha', 'inversion']:
                        row_data[field_name] = self.safe_float(cell_value)
                    else:
                        row_data[field_name] = self.safe_string(cell_value)
            
            # Solo agregar si la fila tiene datos
            if has_data:
                batch.append(row_data)
                total_rows += 1
            
            # Enviar lote cuando esté lleno
            if len(batch) >= batch_size:
                logger.info(f"Enviando lote con {len(batch)} registros (total procesadas: {total_rows} filas)")
                yield batch
                batch = []
        
        # Enviar último lote si tiene datos
        if batch:
            logger.info(f"Enviando último lote con {len(batch)} registros (total procesadas: {total_rows} filas)")
            yield batch
        
        workbook.close()
        logger.info(f"Extracción completada: {total_rows} registros totales")
    
    def get_total_rows(self) -> int:
        """Obtiene el total de filas con datos en el Excel."""
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook[self.sheet_name]
        
        # Contar filas con datos (excluyendo header)
        count = 0
        for row_num in range(2, sheet.max_row + 1):
            # Verificar si la fila tiene al menos un dato
            has_data = False
            for col_num in range(1, min(18, sheet.max_column + 1)):  # Máximo 17 columnas esperadas
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            
            if has_data:
                count += 1
        
        workbook.close()
        return count